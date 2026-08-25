from __future__ import annotations

import io
import logging
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from stages.base import BaseStage, StageResult

logger = logging.getLogger(__name__)

EXPECTED_HH = 6
GAP_THRESHOLD = timedelta(days=2)
HBO_TOLERANCE = timedelta(days=3)
OUTPUT_PATH = Path('Output') / 'havi_round_status.xlsx'
ASPIRATION_MRCCODES = {'64', '66', '70'}


def group_rounds(dates: list[date]) -> list[tuple[date, date]]:
    """Group a list of dates into round windows where gap > 2 days starts a new round.

    Returns list of (first_date, last_date) tuples in chronological order.
    """
    if not dates:
        return []
    sorted_dates = sorted(set(dates))
    rounds: list[tuple[date, date]] = []
    window_start = sorted_dates[0]
    window_end = sorted_dates[0]
    for d in sorted_dates[1:]:
        if d - window_end > GAP_THRESHOLD:
            rounds.append((window_start, window_end))
            window_start = d
        window_end = d
    rounds.append((window_start, window_end))
    return rounds


def hlc_round_counts(df: pd.DataFrame, round_start: date, round_end: date) -> dict:
    """Return per-night, per-clocation HH counts for an HLC round window.

    df must be pre-filtered to one mrccode and datasource=1.
    Columns required: hhid, dateofcollection, clocation.
    """
    coll_dates = pd.to_datetime(df['dateofcollection']).dt.date
    window = df[(coll_dates >= round_start) & (coll_dates <= round_end)].copy()
    window = window.assign(
        _date=pd.to_datetime(window['dateofcollection']).dt.date,
        clocation=pd.to_numeric(window['clocation'], errors='coerce'),
    )

    nights = sorted(window['_date'].unique())
    n1_date = nights[0] if len(nights) >= 1 else None
    n2_date = nights[1] if len(nights) >= 2 else None

    def _hh_count(night: date | None, cloc: int) -> int:
        if night is None:
            return 0
        return int(window[(window['_date'] == night) & (window['clocation'] == cloc)]['hhid'].nunique())

    n1_indoor = _hh_count(n1_date, 2)
    n1_outdoor = _hh_count(n1_date, 1)
    n2_indoor = _hh_count(n2_date, 2)
    n2_outdoor = _hh_count(n2_date, 1)

    complete = (
        n2_date is not None
        and n1_indoor == EXPECTED_HH
        and n1_outdoor == EXPECTED_HH
        and n2_indoor == EXPECTED_HH
        and n2_outdoor == EXPECTED_HH
    )

    return {
        'n1_date': n1_date,
        'n2_date': n2_date,
        'n1_indoor': n1_indoor,
        'n1_outdoor': n1_outdoor,
        'n2_indoor': n2_indoor,
        'n2_outdoor': n2_outdoor,
        'complete': complete,
    }


def hbo_round_counts(df: pd.DataFrame, round_start: date, round_end: date) -> dict:
    """Return HBO HH count aligned to the given HLC round window (±3 days tolerance).

    df must be pre-filtered to one mrccode.
    Columns required: hhid, dateofobservation.
    """
    if df.empty:
        return {'hbo_dates': '', 'hbo_hh': 0, 'hbo_complete': False}

    df = df.assign(_date=pd.to_datetime(df['dateofobservation']).dt.date)
    window = df[
        (df['_date'] >= round_start - HBO_TOLERANCE)
        & (df['_date'] <= round_end + HBO_TOLERANCE)
    ]

    if window.empty:
        return {'hbo_dates': '', 'hbo_hh': 0, 'hbo_complete': False}

    hbo_hh = int(window['hhid'].nunique())
    sorted_dates = sorted(window['_date'].unique())
    if len(sorted_dates) == 1:
        hbo_dates = str(sorted_dates[0])
    else:
        hbo_dates = f"{sorted_dates[0]} / {sorted_dates[-1]}"

    return {
        'hbo_dates': hbo_dates,
        'hbo_hh': hbo_hh,
        'hbo_complete': hbo_hh >= EXPECTED_HH,
    }


def hbo_missing_hh(
    collection_df: pd.DataFrame,
    hbo_df: pd.DataFrame,
    round_start: date,
    round_end: date,
) -> list[dict]:
    """Return hhids that have ento collection in the round but no HBO within tolerance.

    Checks at the per-HH per-round level: an HH with at least one HBO entry
    anywhere in [round_start - tolerance, round_end + tolerance] is not flagged.
    """
    coll_dates = pd.to_datetime(collection_df['dateofcollection']).dt.date
    window = collection_df[
        (coll_dates >= round_start) & (coll_dates <= round_end)
    ].copy()
    window = window.assign(_date=pd.to_datetime(window['dateofcollection']).dt.date)
    collection_hhids = set(window['hhid'].unique())

    if hbo_df.empty:
        return [{'hhid': hh} for hh in sorted(collection_hhids)]

    hbo = hbo_df.assign(_date=pd.to_datetime(hbo_df['dateofobservation']).dt.date)
    hbo_in_window = hbo[
        (hbo['_date'] >= round_start - HBO_TOLERANCE)
        & (hbo['_date'] <= round_end + HBO_TOLERANCE)
    ]
    hbo_hhids = set(hbo_in_window['hhid'].unique())

    return [{'hhid': hh} for hh in sorted(collection_hhids - hbo_hhids)]


def hbo_wrong_dates(
    collection_df: pd.DataFrame,
    hbo_df: pd.DataFrame,
    round_start: date,
    round_end: date,
) -> list[dict]:
    """Return HBO entries within tolerance of the round but not matching any collection night."""
    window = collection_df[
        (pd.to_datetime(collection_df['dateofcollection']).dt.date >= round_start)
        & (pd.to_datetime(collection_df['dateofcollection']).dt.date <= round_end)
    ].copy()
    window = window.assign(_date=pd.to_datetime(window['dateofcollection']).dt.date)
    collection_nights = set(window['_date'].unique())

    if hbo_df.empty:
        return []

    hbo = hbo_df.assign(_date=pd.to_datetime(hbo_df['dateofobservation']).dt.date)
    hbo_in_window = hbo[
        (hbo['_date'] >= round_start - HBO_TOLERANCE)
        & (hbo['_date'] <= round_end + HBO_TOLERANCE)
    ]

    flagged = []
    for _, row in hbo_in_window.iterrows():
        if row['_date'] not in collection_nights:
            flagged.append({'hhid': row['hhid'], 'hbo_date': row['_date']})
    return flagged


def person_count_issues(
    hbo_hh_df: pd.DataFrame,
    hbo_person_df: pd.DataFrame,
) -> list[dict]:
    """Return sessions where person record count does not match numpeople on the HH form."""
    if hbo_hh_df.empty:
        return []

    if hbo_person_df.empty:
        actual_counts = pd.Series(0, index=hbo_hh_df['session_id'], dtype=int).rename('actual')
    else:
        actual_counts = hbo_person_df.groupby('session_id').size().rename('actual')

    issues = []
    for _, row in hbo_hh_df.iterrows():
        session_id = row['session_id']
        try:
            expected = int(row['numpeople'])
        except (ValueError, TypeError):
            continue
        actual = int(actual_counts.get(session_id, 0))
        if actual != expected:
            issues.append({
                'session_id': session_id,
                'hhid': row['hhid'],
                'expected': expected,
                'actual': actual,
            })
    return issues


_SUMMARY_COLS = [
    'mrccode', 'site', 'round', 'hlc_dates',
    'hlc_n1_indoor', 'hlc_n1_outdoor', 'hlc_n2_indoor', 'hlc_n2_outdoor',
    'hlc_complete', 'hbo_dates', 'hbo_hh', 'hbo_complete',
]
_INCOMPLETE_COLS = [
    'mrccode', 'site', 'round', 'collection_type', 'hhid', 'date', 'clocation', 'status',
]
_ASP_COLS = ['mrccode', 'site', 'asp_dates', 'asp_hh']

_HEADER_FILL = PatternFill('solid', fgColor='4F81BD')
_HEADER_FONT = Font(bold=True, color='FFFFFF')
_COMPLETE_FILL = PatternFill('solid', fgColor='C6EFCE')
_INCOMPLETE_FILL = PatternFill('solid', fgColor='FFC7CE')
_BEHIND_FILL = PatternFill('solid', fgColor='FFE699')  # amber — site has fewer rounds than the max


def _write_sheet(ws, cols: list[str], rows: list[dict]) -> None:
    """Write headers + data rows to a worksheet with basic formatting."""
    for col_idx, col in enumerate(cols, 1):
        cell = ws.cell(1, col_idx, col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col) + 2)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(cols, 1):
            val = row.get(col, '')
            cell = ws.cell(row_idx, col_idx, val)
            if col in ('hlc_complete', 'hbo_complete'):
                cell.fill = _COMPLETE_FILL if val else _INCOMPLETE_FILL


def build_round_status_excel(
    summary_rows: list[dict],
    incomplete_rows: list[dict],
    asp_rows: list[dict],
) -> bytes:
    """Build a three-sheet Excel workbook and return raw bytes."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    _write_sheet(ws_summary, _SUMMARY_COLS, summary_rows)

    # Highlight rows for sites that have fewer rounds than the global maximum.
    if summary_rows:
        from collections import defaultdict
        site_max: dict[str, int] = defaultdict(int)
        for row in summary_rows:
            mrc = row['mrccode']
            site_max[mrc] = max(site_max[mrc], row['round'])
        from statistics import mode
        expected_round = mode(site_max.values())  # majority round — sites ahead don't skew the baseline
        behind = {mrc for mrc, r in site_max.items() if r < expected_round}
        if behind:
            mrc_col = _SUMMARY_COLS.index('mrccode') + 1
            complete_cols = {_SUMMARY_COLS.index(c) + 1 for c in ('hlc_complete', 'hbo_complete')}
            for row_idx in range(2, ws_summary.max_row + 1):
                if ws_summary.cell(row_idx, mrc_col).value in behind:
                    for col_idx in range(1, len(_SUMMARY_COLS) + 1):
                        # Don't overwrite the green/red complete cell fills
                        if col_idx not in complete_cols:
                            ws_summary.cell(row_idx, col_idx).fill = _BEHIND_FILL

    ws_asp = wb.create_sheet('Aspirations')
    _write_sheet(ws_asp, _ASP_COLS, asp_rows)

    ws_detail = wb.create_sheet('Incomplete Detail')
    _write_sheet(ws_detail, _INCOMPLETE_COLS, incomplete_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class RoundStatusStage(BaseStage):
    name = 'round_status'
    dependencies: list[str] = []

    def run(self) -> StageResult:
        mrc_sites: dict[str, str] = self.config.get('mrc_sites') or {}

        # ── Load silver tables ────────────────────────────────────────────
        try:
            collection_df = pd.read_sql('SELECT * FROM silver_havi."ento_collection"', self.engine)
        except Exception as exc:
            return StageResult(success=False, errors=[f'Failed to read ento_collection: {exc}'])

        try:
            hbo_df = pd.read_sql('SELECT * FROM silver_havi."hbo_household"', self.engine)
        except Exception as exc:
            return StageResult(success=False, errors=[f'Failed to read hbo_household: {exc}'])

        try:
            person_df = pd.read_sql('SELECT * FROM silver_havi."hbo_person"', self.engine)
        except Exception as exc:
            return StageResult(success=False, errors=[f'Failed to read hbo_person: {exc}'])

        if collection_df.empty:
            return StageResult(success=False, errors=['No HLC data found in silver schema'])

        # Normalise types
        collection_df['dateofcollection'] = pd.to_datetime(collection_df['dateofcollection']).dt.date
        collection_df['datasource'] = collection_df['datasource'].astype(str)
        collection_df['mrccode'] = collection_df['mrccode'].astype(str)
        if not hbo_df.empty:
            hbo_df['dateofobservation'] = pd.to_datetime(hbo_df['dateofobservation']).dt.date
            hbo_df['mrccode'] = hbo_df['mrccode'].astype(str)
        if not person_df.empty:
            # hbo_person has no mrccode — derive it from hbo_household via session_id
            if not hbo_df.empty and 'session_id' in person_df.columns:
                mrc_map = hbo_df[['session_id', 'mrccode']].drop_duplicates('session_id')
                person_df = person_df.merge(mrc_map, on='session_id', how='left')
            if 'mrccode' not in person_df.columns:
                person_df['mrccode'] = ''
            person_df['mrccode'] = person_df['mrccode'].fillna('').astype(str)

        summary_rows: list[dict] = []
        incomplete_rows: list[dict] = []
        asp_rows: list[dict] = []

        for mrccode, site in sorted(mrc_sites.items(), key=lambda x: int(x[0])):
            site_collection = collection_df[collection_df['mrccode'] == mrccode]
            site_hbo = hbo_df[hbo_df['mrccode'] == mrccode] if not hbo_df.empty else pd.DataFrame()
            site_person = person_df[person_df['mrccode'] == mrccode] if not person_df.empty else pd.DataFrame()

            # ── HLC rounds (datasource=1) ────────────────────────────────
            hlc_df = site_collection[site_collection['datasource'] == '1']
            hlc_dates = list(hlc_df['dateofcollection'].dropna().unique())
            hlc_rounds = group_rounds(hlc_dates)

            for round_num, (r_start, r_end) in enumerate(hlc_rounds, 1):
                counts = hlc_round_counts(hlc_df, r_start, r_end)
                hbo_counts = hbo_round_counts(site_hbo, r_start, r_end)

                hlc_dates_str = (
                    f"{r_start} / {r_end}" if r_start != r_end else str(r_start)
                )

                summary_rows.append({
                    'mrccode': mrccode,
                    'site': site,
                    'round': round_num,
                    'hlc_dates': hlc_dates_str,
                    'hlc_n1_indoor': counts['n1_indoor'],
                    'hlc_n1_outdoor': counts['n1_outdoor'],
                    'hlc_n2_indoor': counts['n2_indoor'],
                    'hlc_n2_outdoor': counts['n2_outdoor'],
                    'hlc_complete': counts['complete'],
                    'hbo_dates': hbo_counts['hbo_dates'],
                    'hbo_hh': hbo_counts['hbo_hh'],
                    'hbo_complete': hbo_counts['hbo_complete'],
                })

                # ── Incomplete detail rows ────────────────────────────────
                if not counts['complete']:
                    for night_label, night_date, n_indoor, n_outdoor in [
                        ('night 1', counts['n1_date'], counts['n1_indoor'], counts['n1_outdoor']),
                        ('night 2', counts['n2_date'], counts['n2_indoor'], counts['n2_outdoor']),
                    ]:
                        if night_date is None:
                            incomplete_rows.append({
                                'mrccode': mrccode, 'site': site, 'round': round_num,
                                'collection_type': 'HLC', 'hhid': '', 'date': str(r_end),
                                'clocation': '', 'status': f'missing {night_label} entirely',
                            })
                            continue
                        if n_indoor < EXPECTED_HH:
                            incomplete_rows.append({
                                'mrccode': mrccode, 'site': site, 'round': round_num,
                                'collection_type': 'HLC', 'hhid': '',
                                'date': str(night_date), 'clocation': 'indoor',
                                'status': f'missing {night_label} indoor ({n_indoor}/{EXPECTED_HH} HH)',
                            })
                        if n_outdoor < EXPECTED_HH:
                            incomplete_rows.append({
                                'mrccode': mrccode, 'site': site, 'round': round_num,
                                'collection_type': 'HLC', 'hhid': '',
                                'date': str(night_date), 'clocation': 'outdoor',
                                'status': f'missing {night_label} outdoor ({n_outdoor}/{EXPECTED_HH} HH)',
                            })

                if not hbo_counts['hbo_complete']:
                    incomplete_rows.append({
                        'mrccode': mrccode, 'site': site, 'round': round_num,
                        'collection_type': 'HBO', 'hhid': '',
                        'date': hbo_counts['hbo_dates'], 'clocation': '',
                        'status': f"HBO incomplete ({hbo_counts['hbo_hh']}/{EXPECTED_HH} HH)",
                    })

                # Per-HH missing HBO
                for m in hbo_missing_hh(hlc_df, site_hbo, r_start, r_end):
                    incomplete_rows.append({
                        'mrccode': mrccode, 'site': site, 'round': round_num,
                        'collection_type': 'HBO', 'hhid': m['hhid'],
                        'date': f"{r_start} / {r_end}", 'clocation': '',
                        'status': 'HBO missing for this HH',
                    })

                # Wrong-date HBO
                for w in hbo_wrong_dates(hlc_df, site_hbo, r_start, r_end):
                    incomplete_rows.append({
                        'mrccode': mrccode, 'site': site, 'round': round_num,
                        'collection_type': 'HBO', 'hhid': w['hhid'],
                        'date': str(w['hbo_date']), 'clocation': '',
                        'status': f"wrong date: entered {w['hbo_date']}, expected a collection night in {r_start} / {r_end}",
                    })

            # Person count issues (across all rounds for this site)
            site_hbo_hh = site_hbo if not site_hbo.empty else pd.DataFrame(columns=['session_id', 'hhid', 'numpeople'])
            for p in person_count_issues(site_hbo_hh, site_person):
                session_date = p['session_id'].split('-', 1)[1] if '-' in p['session_id'] else ''
                incomplete_rows.append({
                    'mrccode': mrccode, 'site': site, 'round': '',
                    'collection_type': 'HBO', 'hhid': p['hhid'],
                    'date': session_date, 'clocation': '',
                    'status': f"person count mismatch: {p['actual']} entered, {p['expected']} expected",
                })

            # ── Aspiration visits (datasource=2) ─────────────────────────
            if mrccode in ASPIRATION_MRCCODES:
                asp_df = site_collection[site_collection['datasource'] == '2']
                if not asp_df.empty:
                    asp_dates_list = list(asp_df['dateofcollection'].dropna().unique())
                    for a_start, a_end in group_rounds(asp_dates_list):
                        asp_hh = int(asp_df[
                            (asp_df['dateofcollection'] >= a_start)
                            & (asp_df['dateofcollection'] <= a_end)
                        ]['hhid'].nunique())
                        asp_rows.append({
                            'mrccode': mrccode,
                            'site': site,
                            'asp_dates': f"{a_start} / {a_end}" if a_start != a_end else str(a_start),
                            'asp_hh': asp_hh,
                        })

        # ── Write Excel ───────────────────────────────────────────────────
        xlsx_bytes = build_round_status_excel(summary_rows, incomplete_rows, asp_rows)
        OUTPUT_PATH.parent.mkdir(exist_ok=True)
        OUTPUT_PATH.write_bytes(xlsx_bytes)
        logger.info("Round status report written to %s", OUTPUT_PATH)

        return StageResult(success=True, rows_written=len(summary_rows))
