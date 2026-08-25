from datetime import date
import io
import pytest
import pandas as pd
from openpyxl import load_workbook
from unittest.mock import MagicMock, patch
from stages.round_status import group_rounds, hlc_round_counts, hbo_round_counts, hbo_missing_hh, hbo_wrong_dates, person_count_issues, EXPECTED_HH, build_round_status_excel, RoundStatusStage


def test_empty_dates_returns_empty():
    assert group_rounds([]) == []


def test_single_date_returns_one_window():
    d = date(2026, 5, 5)
    assert group_rounds([d]) == [(d, d)]


def test_two_consecutive_dates_same_round():
    d1, d2 = date(2026, 5, 5), date(2026, 5, 6)
    assert group_rounds([d1, d2]) == [(d1, d2)]


def test_gap_over_two_days_splits_rounds():
    d1, d2 = date(2026, 5, 5), date(2026, 6, 2)
    result = group_rounds([d1, d2])
    assert result == [(d1, d1), (d2, d2)]


def test_two_full_rounds():
    dates = [date(2026, 5, 5), date(2026, 5, 6), date(2026, 6, 2), date(2026, 6, 3)]
    result = group_rounds(dates)
    assert result == [(date(2026, 5, 5), date(2026, 5, 6)), (date(2026, 6, 2), date(2026, 6, 3))]


def test_unsorted_input_sorted_before_grouping():
    dates = [date(2026, 6, 2), date(2026, 5, 5)]
    result = group_rounds(dates)
    assert result[0] == (date(2026, 5, 5), date(2026, 5, 5))
    assert result[1] == (date(2026, 6, 2), date(2026, 6, 2))


def test_gap_of_exactly_two_days_stays_same_round():
    d1, d2 = date(2026, 5, 5), date(2026, 5, 7)
    assert group_rounds([d1, d2]) == [(d1, d2)]


def test_gap_of_three_days_splits_rounds():
    d1, d2 = date(2026, 5, 5), date(2026, 5, 8)
    result = group_rounds([d1, d2])
    assert result == [(d1, d1), (d2, d2)]


def _hlc_row(hhid: str, d: date, clocation: int) -> dict:
    return {'hhid': hhid, 'dateofcollection': d, 'clocation': clocation}


def _full_hlc_df(n1: date, n2: date) -> pd.DataFrame:
    """6 HH × 2 nights × 2 clocations = 24 rows — all complete."""
    rows = []
    for i in range(1, EXPECTED_HH + 1):
        hh = f'HH{i:03}'
        for d in (n1, n2):
            rows.append(_hlc_row(hh, d, 1))  # outdoor
            rows.append(_hlc_row(hh, d, 2))  # indoor
    return pd.DataFrame(rows)


def test_hlc_complete_round():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    df = _full_hlc_df(n1, n2)
    result = hlc_round_counts(df, n1, n2)
    assert result['complete'] is True
    assert result['n1_indoor'] == 6
    assert result['n1_outdoor'] == 6
    assert result['n2_indoor'] == 6
    assert result['n2_outdoor'] == 6
    assert result['n1_date'] == n1
    assert result['n2_date'] == n2


def test_hlc_missing_one_hh_outdoor_night2():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    df = _full_hlc_df(n1, n2)
    # Remove one HH outdoor record on night 2
    df = df[~((df['hhid'] == 'HH001') & (df['dateofcollection'] == n2) & (df['clocation'] == 1))]
    result = hlc_round_counts(df, n1, n2)
    assert result['complete'] is False
    assert result['n2_outdoor'] == 5
    assert result['n2_indoor'] == 6


def test_hlc_single_night_round():
    n1 = date(2026, 6, 17)
    rows = [_hlc_row(f'HH{i:03}', n1, c) for i in range(1, 7) for c in (1, 2)]
    df = pd.DataFrame(rows)
    result = hlc_round_counts(df, n1, n1)
    assert result['n1_date'] == n1
    assert result['n2_date'] is None
    assert result['n1_indoor'] == 6
    assert result['n2_indoor'] == 0
    assert result['complete'] is False  # only 1 night — can't be complete


def test_hlc_empty_dataframe():
    result = hlc_round_counts(pd.DataFrame(columns=['hhid', 'dateofcollection', 'clocation']), date(2026, 6, 18), date(2026, 6, 19))
    assert result['complete'] is False
    assert result['n1_indoor'] == 0


def _hbo_df(dates_per_hh: dict[str, date]) -> pd.DataFrame:
    return pd.DataFrame([
        {'hhid': hh, 'dateofobservation': d}
        for hh, d in dates_per_hh.items()
    ])


def test_hbo_complete_aligned_to_round():
    round_start, round_end = date(2026, 6, 18), date(2026, 6, 19)
    hbo = _hbo_df({f'HH{i:03}': date(2026, 6, 18) for i in range(1, 7)})
    result = hbo_round_counts(hbo, round_start, round_end)
    assert result['hbo_hh'] == 6
    assert result['hbo_complete'] is True


def test_hbo_incomplete_four_hh():
    round_start, round_end = date(2026, 6, 18), date(2026, 6, 19)
    hbo = _hbo_df({f'HH{i:03}': date(2026, 6, 18) for i in range(1, 5)})
    result = hbo_round_counts(hbo, round_start, round_end)
    assert result['hbo_hh'] == 4
    assert result['hbo_complete'] is False


def test_hbo_aligned_within_3_days_tolerance():
    # HBO submitted 3 days after round end — still aligned
    round_start, round_end = date(2026, 6, 18), date(2026, 6, 19)
    hbo = _hbo_df({f'HH{i:03}': date(2026, 6, 22) for i in range(1, 7)})
    result = hbo_round_counts(hbo, round_start, round_end)
    assert result['hbo_hh'] == 6


def test_hbo_outside_tolerance_returns_empty():
    round_start, round_end = date(2026, 6, 18), date(2026, 6, 19)
    # HBO 10 days later — should not align to this round
    hbo = _hbo_df({f'HH{i:03}': date(2026, 6, 29) for i in range(1, 7)})
    result = hbo_round_counts(hbo, round_start, round_end)
    assert result['hbo_hh'] == 0
    assert result['hbo_complete'] is False
    assert result['hbo_dates'] == ''


def test_hbo_empty_dataframe():
    result = hbo_round_counts(pd.DataFrame(columns=['hhid', 'dateofobservation']), date(2026, 6, 18), date(2026, 6, 19))
    assert result['hbo_hh'] == 0
    assert result['hbo_complete'] is False


def _load_wb(data: bytes):
    return load_workbook(io.BytesIO(data))


def test_excel_has_three_sheets():
    wb = _load_wb(build_round_status_excel([], [], []))
    assert wb.sheetnames == ['Summary', 'Aspirations', 'Incomplete Detail']


def test_excel_summary_headers():
    wb = _load_wb(build_round_status_excel([], [], []))
    ws = wb['Summary']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert 'mrccode' in headers
    assert 'hlc_complete' in headers
    assert 'hbo_complete' in headers


def test_excel_summary_row_written():
    row = {
        'mrccode': '12', 'site': 'Kyatiri', 'round': 1,
        'hlc_dates': '2026-05-19 / 2026-05-20',
        'hlc_n1_indoor': 6, 'hlc_n1_outdoor': 6,
        'hlc_n2_indoor': 6, 'hlc_n2_outdoor': 6,
        'hlc_complete': True,
        'hbo_dates': '2026-05-19', 'hbo_hh': 6, 'hbo_complete': True,
    }
    wb = _load_wb(build_round_status_excel([row], [], []))
    ws = wb['Summary']
    assert ws.max_row == 2  # header + 1 data row
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    mrc_col = headers.index('mrccode') + 1
    assert ws.cell(2, mrc_col).value == '12'


def test_excel_incomplete_detail_headers():
    wb = _load_wb(build_round_status_excel([], [], []))
    ws = wb['Incomplete Detail']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert 'status' in headers
    assert 'clocation' in headers


def test_excel_aspiration_row_written():
    asp = {'mrccode': '64', 'site': 'Busitema', 'asp_dates': '2026-05-22 / 2026-05-23', 'asp_hh': 6}
    wb = _load_wb(build_round_status_excel([], [], [asp]))
    ws = wb['Aspirations']
    assert ws.max_row == 2


def _make_stage(collection_rows, hbo_rows, person_rows=None, mrc_sites=None):
    """Build a RoundStatusStage with mocked DB returning given rows."""
    config = MagicMock()
    config.get.side_effect = lambda key, default=None: (
        mrc_sites or {'12': 'Kyatiri', '70': 'Nagongera'}
    ) if key == 'mrc_sites' else default

    engine = MagicMock()

    collection_df = pd.DataFrame(collection_rows) if collection_rows else pd.DataFrame(
        columns=['hhid', 'mrccode', 'dateofcollection', 'clocation', 'datasource']
    )
    hbo_df = pd.DataFrame(hbo_rows) if hbo_rows else pd.DataFrame(
        columns=['hhid', 'mrccode', 'dateofobservation', 'session_id', 'numpeople']
    )
    person_df = pd.DataFrame(person_rows) if person_rows else pd.DataFrame(
        columns=['hhid', 'mrccode', 'session_id']
    )

    with patch('stages.round_status.pd.read_sql', side_effect=[collection_df, hbo_df, person_df]):
        stage = RoundStatusStage(config=config, engine=engine)
        with patch('stages.round_status.build_round_status_excel', return_value=b'XLSX') as mock_excel:
            with patch('stages.round_status.OUTPUT_PATH') as mock_path:
                mock_path.parent.mkdir = MagicMock()
                mock_path.write_bytes = MagicMock()
                result = stage.run()
    return result, mock_excel


def test_stage_returns_failure_on_empty_collection():
    result, _ = _make_stage([], [])
    assert result.success is False
    assert 'No HLC data' in result.errors[0]


def test_stage_writes_output_and_returns_success():
    n1, n2 = date(2026, 5, 19), date(2026, 5, 20)
    rows = []
    for i in range(1, 7):
        hh = f'337010{i:03}'
        for d in (n1, n2):
            rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 1, 'datasource': '1'})
            rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 2, 'datasource': '1'})

    hbo_rows = [
        {'hhid': f'337010{i:03}', 'mrccode': '12', 'dateofobservation': n1,
         'session_id': f'337010{i:03}-2026-05-19', 'numpeople': '3'}
        for i in range(1, 7)
    ]
    person_rows = [
        {'hhid': f'337010{i:03}', 'mrccode': '12', 'session_id': f'337010{i:03}-2026-05-19'}
        for i in range(1, 7)
        for _ in range(3)
    ]

    result, mock_excel = _make_stage(rows, hbo_rows, person_rows)
    assert result.success is True
    assert result.rows_written >= 1
    mock_excel.assert_called_once()


# ── Gap 1: per-HH HBO missing detection ──────────────────────────────────────

def _collection_df_for_round(n1: date, n2: date, hhids: list[str]) -> pd.DataFrame:
    rows = []
    for hh in hhids:
        for d in (n1, n2):
            rows.append({'hhid': hh, 'dateofcollection': d, 'clocation': 1})
            rows.append({'hhid': hh, 'dateofcollection': d, 'clocation': 2})
    return pd.DataFrame(rows)


def test_hbo_missing_hh_returns_empty_when_all_present():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    hhids = [f'HH{i:03}' for i in range(1, 7)]
    coll = _collection_df_for_round(n1, n2, hhids)
    hbo = pd.DataFrame([{'hhid': hh, 'dateofobservation': n1} for hh in hhids]
                       + [{'hhid': hh, 'dateofobservation': n2} for hh in hhids])
    assert hbo_missing_hh(coll, hbo, n1, n2) == []


def test_hbo_missing_hh_flags_hhid_with_no_hbo_in_round():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    hhids = [f'HH{i:03}' for i in range(1, 7)]
    coll = _collection_df_for_round(n1, n2, hhids)
    # HH006 has no HBO at all
    hbo = pd.DataFrame([{'hhid': hh, 'dateofobservation': n1} for hh in hhids[:-1]])
    missing = hbo_missing_hh(coll, hbo, n1, n2)
    assert any(m['hhid'] == 'HH006' for m in missing)


def test_hbo_missing_hh_does_not_flag_hhid_with_partial_hbo():
    # Per-round check: an HH with HBO on at least one night is not flagged as missing
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    hhids = [f'HH{i:03}' for i in range(1, 7)]
    coll = _collection_df_for_round(n1, n2, hhids)
    # HH001 has HBO only on n1 — has SOME HBO for the round, not flagged
    hbo = pd.DataFrame(
        [{'hhid': hh, 'dateofobservation': n1} for hh in hhids]
        + [{'hhid': hh, 'dateofobservation': n2} for hh in hhids[1:]]
    )
    missing = hbo_missing_hh(coll, hbo, n1, n2)
    assert not any(m['hhid'] == 'HH001' for m in missing)


def test_hbo_missing_hh_tolerates_hbo_date_within_3_days():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    hhids = [f'HH{i:03}' for i in range(1, 7)]
    coll = _collection_df_for_round(n1, n2, hhids)
    # All HBO submitted 2 days after round end — still within tolerance
    hbo = pd.DataFrame([{'hhid': hh, 'dateofobservation': date(2026, 6, 21)} for hh in hhids])
    missing = hbo_missing_hh(coll, hbo, n1, n2)
    assert missing == []


def test_stage_incomplete_rows_include_per_hh_missing_hbo():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    coll_rows = []
    for i in range(1, 7):
        hh = f'337010{i:03}'
        for d in (n1, n2):
            coll_rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 1, 'datasource': '1'})
            coll_rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 2, 'datasource': '1'})
    # HH006 has no HBO
    hbo_rows = [
        {'hhid': f'337010{i:03}', 'mrccode': '12', 'dateofobservation': n1,
         'session_id': f'337010{i:03}-2026-06-18', 'numpeople': '2'}
        for i in range(1, 6)
    ]
    person_rows = [
        {'hhid': f'337010{i:03}', 'mrccode': '12', 'session_id': f'337010{i:03}-2026-06-18'}
        for i in range(1, 6) for _ in range(2)
    ]
    _, mock_excel = _make_stage(coll_rows, hbo_rows, person_rows, mrc_sites={'12': 'Kyatiri'})
    _, incomplete_rows, _ = mock_excel.call_args[0]
    hbo_missing = [r for r in incomplete_rows if '337010006' in r.get('hhid', '') and r.get('collection_type') == 'HBO']
    assert len(hbo_missing) >= 1


# ── Gap 2: wrong-date HBO detection ──────────────────────────────────────────

def test_hbo_wrong_dates_returns_empty_when_dates_match_collection_nights():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    hhids = [f'HH{i:03}' for i in range(1, 7)]
    coll = _collection_df_for_round(n1, n2, hhids)
    hbo = pd.DataFrame(
        [{'hhid': hh, 'dateofobservation': n1} for hh in hhids]
        + [{'hhid': hh, 'dateofobservation': n2} for hh in hhids]
    )
    assert hbo_wrong_dates(coll, hbo, n1, n2) == []


def test_hbo_wrong_dates_flags_date_within_tolerance_but_not_a_collection_night():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    hhids = [f'HH{i:03}' for i in range(1, 7)]
    coll = _collection_df_for_round(n1, n2, hhids)
    wrong = date(2026, 6, 20)  # within ±3 tolerance but not a collection night
    hbo = pd.DataFrame(
        [{'hhid': 'HH001', 'dateofobservation': wrong}]
        + [{'hhid': hh, 'dateofobservation': n1} for hh in hhids[1:]]
    )
    flagged = hbo_wrong_dates(coll, hbo, n1, n2)
    assert len(flagged) == 1
    assert flagged[0]['hhid'] == 'HH001'
    assert flagged[0]['hbo_date'] == wrong


def test_hbo_wrong_dates_does_not_flag_dates_outside_tolerance():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    hhids = [f'HH{i:03}' for i in range(1, 7)]
    coll = _collection_df_for_round(n1, n2, hhids)
    # Jun-25 is 6 days after round end — outside tolerance entirely, not counted
    hbo = pd.DataFrame([{'hhid': hh, 'dateofobservation': date(2026, 6, 25)} for hh in hhids])
    assert hbo_wrong_dates(coll, hbo, n1, n2) == []


def test_stage_incomplete_rows_include_wrong_date_hbo():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    coll_rows = []
    for i in range(1, 7):
        hh = f'337010{i:03}'
        for d in (n1, n2):
            coll_rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 1, 'datasource': '1'})
            coll_rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 2, 'datasource': '1'})
    wrong_date = date(2026, 6, 20)
    hbo_rows = [
        {'hhid': '337010001', 'mrccode': '12', 'dateofobservation': wrong_date,
         'session_id': '337010001-2026-06-20', 'numpeople': '2'},
        *[
            {'hhid': f'337010{i:03}', 'mrccode': '12', 'dateofobservation': n1,
             'session_id': f'337010{i:03}-2026-06-18', 'numpeople': '2'}
            for i in range(2, 7)
        ],
    ]
    person_rows = [
        {'hhid': f'337010{i:03}', 'mrccode': '12', 'session_id': f'337010{i:03}-2026-06-18'}
        for i in range(2, 7) for _ in range(2)
    ] + [
        {'hhid': '337010001', 'mrccode': '12', 'session_id': '337010001-2026-06-20'}
        for _ in range(2)
    ]
    _, mock_excel = _make_stage(coll_rows, hbo_rows, person_rows, mrc_sites={'12': 'Kyatiri'})
    _, incomplete_rows, _ = mock_excel.call_args[0]
    wrong = [r for r in incomplete_rows if 'wrong date' in r.get('status', '')]
    assert len(wrong) >= 1
    assert '337010001' in wrong[0]['hhid']


# ── Gap 3: person count mismatch ─────────────────────────────────────────────

def test_person_count_issues_returns_empty_when_counts_match():
    hbo_hh = pd.DataFrame([
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001', 'numpeople': '3'},
    ])
    hbo_person = pd.DataFrame([
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
    ])
    assert person_count_issues(hbo_hh, hbo_person) == []


def test_person_count_issues_flags_when_actual_less_than_numpeople():
    hbo_hh = pd.DataFrame([
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001', 'numpeople': '4'},
    ])
    hbo_person = pd.DataFrame([
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
    ])
    issues = person_count_issues(hbo_hh, hbo_person)
    assert len(issues) == 1
    assert issues[0]['session_id'] == 'HH001-2026-06-18'
    assert issues[0]['expected'] == 4
    assert issues[0]['actual'] == 2


def test_person_count_issues_flags_when_zero_persons_entered():
    hbo_hh = pd.DataFrame([
        {'session_id': 'HH001-2026-06-03', 'hhid': 'HH001', 'numpeople': '4'},
    ])
    hbo_person = pd.DataFrame(columns=['session_id', 'hhid'])
    issues = person_count_issues(hbo_hh, hbo_person)
    assert len(issues) == 1
    assert issues[0]['actual'] == 0
    assert issues[0]['expected'] == 4


def test_person_count_issues_flags_when_actual_more_than_numpeople():
    hbo_hh = pd.DataFrame([
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001', 'numpeople': '2'},
    ])
    hbo_person = pd.DataFrame([
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
        {'session_id': 'HH001-2026-06-18', 'hhid': 'HH001'},
    ])
    issues = person_count_issues(hbo_hh, hbo_person)
    assert len(issues) == 1
    assert issues[0]['expected'] == 2
    assert issues[0]['actual'] == 4


def test_stage_incomplete_rows_include_person_count_mismatch():
    n1, n2 = date(2026, 6, 18), date(2026, 6, 19)
    coll_rows = []
    for i in range(1, 7):
        hh = f'337010{i:03}'
        for d in (n1, n2):
            coll_rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 1, 'datasource': '1'})
            coll_rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 2, 'datasource': '1'})
    hbo_rows = [
        {'hhid': f'337010{i:03}', 'mrccode': '12', 'dateofobservation': n1,
         'session_id': f'337010{i:03}-2026-06-18', 'numpeople': '3'}
        for i in range(1, 7)
    ]
    # HH001 only has 1 person instead of 3
    person_rows = (
        [{'hhid': '337010001', 'mrccode': '12', 'session_id': '337010001-2026-06-18'}]
        + [
            {'hhid': f'337010{i:03}', 'mrccode': '12', 'session_id': f'337010{i:03}-2026-06-18'}
            for i in range(2, 7) for _ in range(3)
        ]
    )
    _, mock_excel = _make_stage(coll_rows, hbo_rows, person_rows, mrc_sites={'12': 'Kyatiri'})
    _, incomplete_rows, _ = mock_excel.call_args[0]
    person_issues = [r for r in incomplete_rows if 'person count' in r.get('status', '')]
    assert len(person_issues) >= 1
    assert '337010001' in person_issues[0]['hhid']


def test_stage_does_not_crash_when_hbo_empty_but_person_has_rows():
    """hbo_person has no native mrccode column; if hbo_household is empty the
    stage must not blow up deriving one (regression for KeyError: 'mrccode')."""
    n1, n2 = date(2026, 5, 19), date(2026, 5, 20)
    coll_rows = []
    for i in range(1, 7):
        hh = f'337010{i:03}'
        for d in (n1, n2):
            coll_rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 1, 'datasource': '1'})
            coll_rows.append({'hhid': hh, 'mrccode': '12', 'dateofcollection': d, 'clocation': 2, 'datasource': '1'})
    person_rows = [{'hhid': '337010001', 'session_id': '337010001-2026-05-19'}]

    result, _ = _make_stage(coll_rows, [], person_rows, mrc_sites={'12': 'Kyatiri'})
    assert result.success is True
