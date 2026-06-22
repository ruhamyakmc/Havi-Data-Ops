from __future__ import annotations

import smtplib
from datetime import datetime, timezone, timedelta
from unittest.mock import MagicMock, patch

# A Monday in EAT — used to keep field-report tests independent of the day they run
_EAT = timezone(timedelta(hours=3))
_MONDAY_EAT = datetime(2026, 6, 1, 9, 0, 0, tzinfo=_EAT)  # weekday() == 0
_TUESDAY_EAT = datetime(2026, 6, 2, 9, 0, 0, tzinfo=_EAT)  # weekday() == 1

import pandas as pd
import pytest
from cryptography.fernet import Fernet

import io as _io

from openpyxl import load_workbook

from modules.notifier import (
    _build_stage_summary,
    _build_validation_summary,
    _build_validation_details_excel,
    _load_smtp_password,
    _query_validation_report,
    _send,
    send_pipeline_report,
)
from stages.base import StageResult


def test_load_smtp_password_roundtrip(tmp_path):
    key = Fernet.generate_key()
    cipher = Fernet(key)
    key_file = tmp_path / 'smtp.key'
    ini_file = tmp_path / 'smtp.ini'
    key_file.write_text(key.decode())
    ini_file.write_text(f"Password={cipher.encrypt(b's3cr3t').decode()}\n")

    assert _load_smtp_password(str(ini_file), str(key_file)) == 's3cr3t'


def test_load_smtp_password_missing_raises(tmp_path):
    key = Fernet.generate_key()
    key_file = tmp_path / 'smtp.key'
    ini_file = tmp_path / 'smtp.ini'
    key_file.write_text(key.decode())
    ini_file.write_text('Host=smtp.example.com\n')

    with pytest.raises(KeyError, match='Password'):
        _load_smtp_password(str(ini_file), str(key_file))


def test_query_validation_report_reads_gold_havi():
    with patch('pandas.read_sql', return_value=pd.DataFrame()) as mock_read:
        _query_validation_report(MagicMock())
    assert 'gold_havi.ds_validation_report' in mock_read.call_args[0][0]


def test_query_validation_report_returns_none_on_db_error():
    with patch('pandas.read_sql', side_effect=Exception('connection refused')):
        result = _query_validation_report(MagicMock())
    assert result is None


def test_build_stage_summary_shows_all_statuses():
    results = {
        'sqlite_to_bronze': StageResult(success=True, rows_written=5416),
        'bronze_to_silver': StageResult(success=False, errors=['err']),
    }
    stages = ['sqlite_to_bronze', 'bronze_to_silver', 'transform_havi']
    text = _build_stage_summary(results, stages)
    assert 'OK' in text and 'sqlite_to_bronze' in text
    assert 'FAIL' in text and 'bronze_to_silver' in text
    assert 'SKIP' in text and 'transform_havi' in text
    assert '5,416' in text


def test_build_stage_summary_no_rows_for_zero():
    results = {'transform_havi': StageResult(success=True, rows_written=0)}
    text = _build_stage_summary(results, ['transform_havi'])
    assert 'OK' in text
    assert '0' not in text


def test_build_validation_summary_none_returns_unavailable():
    text = _build_validation_summary(None)
    assert 'unavailable' in text.lower()
    assert 'measures_havi' in text


def test_build_validation_summary_groups_by_severity_country_site():
    report = pd.DataFrame({
        'severity': ['ERROR', 'WARNING', 'WARNING'],
        'check': ['dup_id', 'missing_child_records', 'sparse_col'],
        'country': ['Uganda', 'Uganda', 'Uganda'],
        'site': ['Mbarara', 'Mbarara', 'Bushenyi'],
        'record_count': [2, 3, 1],
    })
    text = _build_validation_summary(report)
    assert 'ERRORS' in text
    assert 'WARNINGS' in text
    assert 'Uganda / Mbarara' in text
    assert 'Uganda / Bushenyi' in text
    assert 'dup_id' in text
    assert 'see attachment' in text.lower()


def test_build_validation_summary_empty_df_returns_header_only():
    report = pd.DataFrame(columns=['severity', 'check', 'country', 'site', 'record_count'])
    text = _build_validation_summary(report)
    assert 'ERRORS' not in text
    assert 'WARNINGS' not in text


def _make_email_cfg(tmp_path):
    key = Fernet.generate_key()
    cipher = Fernet(key)
    key_file = tmp_path / 'smtp.key'
    ini_file = tmp_path / 'smtp.ini'
    key_file.write_text(key.decode())
    ini_file.write_text(f"Password={cipher.encrypt(b's3cr3t').decode()}\n")
    return {
        'smtp_host': 'smtp.example.com',
        'smtp_port': 587,
        'sender': 'havi@example.com',
        'smtp_username': 'user@example.com',
        'pipeline_recipients': ['admin@example.com'],
        'field_cc': ['cc@example.com'],
        'keyfiles': {
            'smtp_ini': str(ini_file),
            'smtp_key': str(key_file),
        },
    }


def _config(email_cfg):
    return {'email': email_cfg}


def test_send_pipeline_report_no_config_is_silent():
    send_pipeline_report(
        results={'sqlite_to_bronze': StageResult(success=False)},
        stages=['sqlite_to_bronze'],
        engine=MagicMock(),
        config={},
    )


def test_send_pipeline_report_always_sends_to_pipeline_recipients(tmp_path):
    config = _config(_make_email_cfg(tmp_path))
    results = {'sqlite_to_bronze': StageResult(success=True)}
    clean_report = pd.DataFrame(columns=['severity', 'check', 'country', 'site'])

    mock_smtp_instance = MagicMock()
    with patch('modules.notifier._query_validation_report', return_value=clean_report):
        with patch('smtplib.SMTP') as mock_smtp_cls:
            mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
            send_pipeline_report(
                results=results, stages=['sqlite_to_bronze'],
                engine=MagicMock(), config=config,
            )

    mock_smtp_instance.sendmail.assert_called_once()
    assert mock_smtp_instance.sendmail.call_args[0][1] == ['admin@example.com']


def test_send_pipeline_report_failed_subject_says_failed(tmp_path):
    config = _config(_make_email_cfg(tmp_path))
    results = {'sqlite_to_bronze': StageResult(success=False, errors=['boom'])}

    mock_smtp_instance = MagicMock()
    with patch('modules.notifier._query_validation_report', return_value=None):
        with patch('smtplib.SMTP') as mock_smtp_cls:
            mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
            send_pipeline_report(
                results=results, stages=['sqlite_to_bronze'],
                engine=MagicMock(), config=config,
            )

    msg_string = mock_smtp_instance.sendmail.call_args[0][2]
    assert 'FAILED' in msg_string
    assert 'HAVI Pipeline' in msg_string


def test_send_pipeline_report_field_email_not_sent_without_clusters(tmp_path):
    # When config has no clusters block, no field emails are sent — only pipeline status.
    config = _config(_make_email_cfg(tmp_path))
    results = {'sqlite_to_bronze': StageResult(success=True)}
    report = pd.DataFrame({
        'severity': ['WARNING'], 'check': ['sparse_column'],
        'country': ['Uganda'], 'site': ['Mbarara'],
        'mrccode': ['25'],
        'record_count': [2],
    })
    sendmail_calls = []
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = lambda *a, **kw: sendmail_calls.append(a)

    with patch('modules.notifier._query_validation_report', return_value=report):
        with patch('smtplib.SMTP') as mock_smtp_cls:
            mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
            send_pipeline_report(
                results=results, stages=['sqlite_to_bronze'],
                engine=MagicMock(), config=config,
            )

    # Only the pipeline status email is sent
    assert len(sendmail_calls) == 1
    assert sendmail_calls[0][1] == ['admin@example.com']


def test_send_pipeline_report_does_not_raise_on_smtp_error(tmp_path):
    config = _config(_make_email_cfg(tmp_path))
    results = {'sqlite_to_bronze': StageResult(success=False)}

    with patch('modules.notifier._query_validation_report', return_value=None):
        with patch('smtplib.SMTP', side_effect=smtplib.SMTPException('conn refused')):
            send_pipeline_report(
                results=results, stages=['sqlite_to_bronze'],
                engine=MagicMock(), config=config,
            )


def _make_smtp_cfg(tmp_path):
    """Minimal email config for testing _send directly."""
    key = Fernet.generate_key()
    cipher = Fernet(key)
    key_file = tmp_path / 'smtp.key'
    ini_file = tmp_path / 'smtp.ini'
    key_file.write_text(key.decode())
    ini_file.write_text(f"Password={cipher.encrypt(b's3cr3t').decode()}\n")
    return {
        'smtp_host': 'smtp.example.com',
        'smtp_port': 587,
        'sender': 'havi@example.com',
        'smtp_username': 'user@example.com',
        'keyfiles': {'smtp_ini': str(ini_file), 'smtp_key': str(key_file)},
    }


def test_send_includes_cc_header_when_cc_provided(tmp_path):
    email_cfg = _make_smtp_cfg(tmp_path)
    captured = {}
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = (
        lambda sender, recipients, msg: captured.update({'recipients': recipients, 'msg': msg})
    )
    with patch('smtplib.SMTP') as mock_smtp_cls:
        mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
        _send(email_cfg, ['to@example.com'], 'Subject', 'plain', '<b>html</b>',
              cc=['cc1@example.com', 'cc2@example.com'])

    assert 'Cc: cc1@example.com, cc2@example.com' in captured['msg']
    assert 'cc1@example.com' in captured['recipients']
    assert 'cc2@example.com' in captured['recipients']


def test_send_no_cc_does_not_add_cc_header(tmp_path):
    email_cfg = _make_smtp_cfg(tmp_path)
    captured = {}
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = (
        lambda sender, recipients, msg: captured.update({'msg': msg})
    )
    with patch('smtplib.SMTP') as mock_smtp_cls:
        mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
        _send(email_cfg, ['to@example.com'], 'Subject', 'plain', '<b>html</b>')

    assert 'Cc:' not in captured['msg']


def _make_cluster_config(tmp_path):
    """Full config with clusters and field_cc for cluster routing tests."""
    key = Fernet.generate_key()
    cipher = Fernet(key)
    key_file = tmp_path / 'smtp.key'
    ini_file = tmp_path / 'smtp.ini'
    key_file.write_text(key.decode())
    ini_file.write_text(f"Password={cipher.encrypt(b's3cr3t').decode()}\n")
    return {
        'email': {
            'smtp_host': 'smtp.example.com',
            'smtp_port': 587,
            'sender': 'havi@example.com',
            'smtp_username': 'user@example.com',
            'pipeline_recipients': ['admin@example.com'],
            'field_cc': ['cc@example.com'],
            'keyfiles': {'smtp_ini': str(ini_file), 'smtp_key': str(key_file)},
        },
        'clusters': {
            'A': {
                'supervisor': 'Sisye Paul',
                'to': ['supervisorA@example.com'],
                'mrc_codes': ['23', '25'],
            },
            'B': {
                'supervisor': 'Otto Geoffrey',
                'to': ['supervisorB@example.com'],
                'mrc_codes': ['12'],
            },
        },
        'mrc_sites': {
            '23': 'Atiak HCIV - Amuru',
            '25': 'Awach HCIV - Gulu',
            '12': 'Kyatiri HCIII - Masindi',
        },
    }


def test_cluster_routing_sends_to_correct_supervisor(tmp_path):
    config = _make_cluster_config(tmp_path)
    # Only Cluster A (mrccode 23) has an issue
    report = pd.DataFrame({
        'severity': ['ERROR'],
        'check': ['count_mismatch'],
        'country': ['Uganda'],
        'site': ['Atiak HCIV - Amuru'],
        'mrccode': ['23'],
        'record_count': [1],
    })
    sendmail_calls = []
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = lambda *a, **kw: sendmail_calls.append(a)

    with patch('modules.notifier._query_validation_report', return_value=report):
        with patch('modules.notifier._build_validation_details_excel', return_value=b'xlsx'):
            with patch('modules.notifier.datetime') as mock_dt:
                mock_dt.now.return_value = _MONDAY_EAT
                with patch('smtplib.SMTP') as mock_smtp_cls:
                    mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
                    send_pipeline_report(
                        results={'sqlite_to_bronze': StageResult(success=True)},
                        stages=['sqlite_to_bronze'],
                        engine=MagicMock(),
                        config=config,
                    )

    all_recipients = [r for call in sendmail_calls for r in call[1]]
    assert 'supervisorA@example.com' in all_recipients
    assert 'supervisorB@example.com' not in all_recipients


def test_cluster_routing_cc_recipients_receive_all_cluster_emails(tmp_path):
    config = _make_cluster_config(tmp_path)
    report = pd.DataFrame({
        'severity': ['ERROR'],
        'check': ['count_mismatch'],
        'country': ['Uganda'],
        'site': ['Atiak HCIV - Amuru'],
        'mrccode': ['23'],
        'record_count': [1],
    })
    sendmail_calls = []
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = lambda *a, **kw: sendmail_calls.append(a)

    with patch('modules.notifier._query_validation_report', return_value=report):
        with patch('modules.notifier._build_validation_details_excel', return_value=b'xlsx'):
            with patch('modules.notifier.datetime') as mock_dt:
                mock_dt.now.return_value = _MONDAY_EAT
                with patch('smtplib.SMTP') as mock_smtp_cls:
                    mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
                    send_pipeline_report(
                        results={'sqlite_to_bronze': StageResult(success=True)},
                        stages=['sqlite_to_bronze'],
                        engine=MagicMock(),
                        config=config,
                    )

    # cc@example.com should appear in the cluster A field email recipients
    field_call = next(
        c for c in sendmail_calls if 'supervisorA@example.com' in c[1]
    )
    assert 'cc@example.com' in field_call[1]


def test_cluster_field_email_subject_contains_cluster_and_supervisor(tmp_path):
    config = _make_cluster_config(tmp_path)
    report = pd.DataFrame({
        'severity': ['WARNING'],
        'check': ['sparse_col'],
        'country': ['Uganda'],
        'site': ['Atiak HCIV - Amuru'],
        'mrccode': ['23'],
        'record_count': [1],
    })
    sendmail_calls = []
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = lambda *a, **kw: sendmail_calls.append(a)

    with patch('modules.notifier._query_validation_report', return_value=report):
        with patch('modules.notifier._build_validation_details_excel', return_value=b'xlsx'):
            with patch('modules.notifier.datetime') as mock_dt:
                mock_dt.now.return_value = _MONDAY_EAT
                with patch('smtplib.SMTP') as mock_smtp_cls:
                    mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
                    send_pipeline_report(
                        results={'sqlite_to_bronze': StageResult(success=True)},
                        stages=['sqlite_to_bronze'],
                        engine=MagicMock(),
                        config=config,
                    )

    field_call = next(c for c in sendmail_calls if 'supervisorA@example.com' in c[1])
    assert 'Cluster A' in field_call[2]
    assert 'Sisye Paul' in field_call[2]


def test_cluster_field_email_body_does_not_contain_stage_summary(tmp_path):
    config = _make_cluster_config(tmp_path)
    report = pd.DataFrame({
        'severity': ['ERROR'],
        'check': ['count_mismatch'],
        'country': ['Uganda'],
        'site': ['Atiak HCIV - Amuru'],
        'mrccode': ['23'],
        'record_count': [1],
    })
    sendmail_calls = []
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = lambda *a, **kw: sendmail_calls.append(a)

    with patch('modules.notifier._query_validation_report', return_value=report):
        with patch('modules.notifier._build_validation_details_excel', return_value=b'xlsx'):
            with patch('modules.notifier.datetime') as mock_dt:
                mock_dt.now.return_value = _MONDAY_EAT
                with patch('smtplib.SMTP') as mock_smtp_cls:
                    mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
                    send_pipeline_report(
                        results={'sqlite_to_bronze': StageResult(success=True)},
                        stages=['sqlite_to_bronze'],
                        engine=MagicMock(),
                        config=config,
                    )

    field_call = next(c for c in sendmail_calls if 'supervisorA@example.com' in c[1])
    # Stage summary lines contain stage names like sqlite_to_bronze
    assert 'sqlite_to_bronze' not in field_call[2]


def test_field_reports_not_sent_when_pipeline_has_failures(tmp_path):
    config = _make_cluster_config(tmp_path)
    report = pd.DataFrame({
        'severity': ['ERROR'],
        'check': ['count_mismatch'],
        'country': ['Uganda'],
        'site': ['Atiak HCIV - Amuru'],
        'mrccode': ['23'],
        'record_count': [1],
    })
    sendmail_calls = []
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = lambda *a, **kw: sendmail_calls.append(a)

    with patch('modules.notifier._query_validation_report', return_value=report):
        with patch('modules.notifier._build_validation_details_excel', return_value=b'xlsx'):
            with patch('modules.notifier.datetime') as mock_dt:
                mock_dt.now.return_value = _MONDAY_EAT  # a send-day, so only the failure guard suppresses
                with patch('smtplib.SMTP') as mock_smtp_cls:
                    mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
                    send_pipeline_report(
                        results={'sqlite_to_bronze': StageResult(success=False, errors=['boom'])},
                        stages=['sqlite_to_bronze'],
                        engine=MagicMock(),
                        config=config,
                    )

    # Only the pipeline status email should be sent — no field/cluster reports
    assert len(sendmail_calls) == 1
    all_recipients = sendmail_calls[0][1]
    assert 'supervisorA@example.com' not in all_recipients


def test_field_reports_not_sent_on_non_send_day(tmp_path):
    config = _make_cluster_config(tmp_path)
    report = pd.DataFrame({
        'severity': ['ERROR'],
        'check': ['count_mismatch'],
        'country': ['Uganda'],
        'site': ['Atiak HCIV - Amuru'],
        'mrccode': ['23'],
        'record_count': [1],
    })
    sendmail_calls = []
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = lambda *a, **kw: sendmail_calls.append(a)

    with patch('modules.notifier._query_validation_report', return_value=report):
        with patch('modules.notifier._build_validation_details_excel', return_value=b'xlsx'):
            with patch('modules.notifier.datetime') as mock_dt:
                mock_dt.now.return_value = _TUESDAY_EAT  # not a send-day
                with patch('smtplib.SMTP') as mock_smtp_cls:
                    mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
                    send_pipeline_report(
                        results={'sqlite_to_bronze': StageResult(success=True)},
                        stages=['sqlite_to_bronze'],
                        engine=MagicMock(),
                        config=config,
                    )

    # Only the pipeline status email; no field reports
    assert len(sendmail_calls) == 1
    assert 'supervisorA@example.com' not in sendmail_calls[0][1]


def test_field_reports_sent_on_send_day(tmp_path):
    config = _make_cluster_config(tmp_path)
    report = pd.DataFrame({
        'severity': ['ERROR'],
        'check': ['count_mismatch'],
        'country': ['Uganda'],
        'site': ['Atiak HCIV - Amuru'],
        'mrccode': ['23'],
        'record_count': [1],
    })
    sendmail_calls = []
    mock_smtp_instance = MagicMock()
    mock_smtp_instance.sendmail.side_effect = lambda *a, **kw: sendmail_calls.append(a)

    with patch('modules.notifier._query_validation_report', return_value=report):
        with patch('modules.notifier._build_validation_details_excel', return_value=b'xlsx'):
            with patch('modules.notifier.datetime') as mock_dt:
                mock_dt.now.return_value = _MONDAY_EAT
                with patch('smtplib.SMTP') as mock_smtp_cls:
                    mock_smtp_cls.return_value.__enter__.return_value = mock_smtp_instance
                    send_pipeline_report(
                        results={'sqlite_to_bronze': StageResult(success=True)},
                        stages=['sqlite_to_bronze'],
                        engine=MagicMock(),
                        config=config,
                    )

    all_recipients = [r for call in sendmail_calls for r in call[1]]
    assert 'supervisorA@example.com' in all_recipients


def test_build_validation_details_excel_creates_sheet_for_every_check():
    """Every distinct check in the report must produce its own sheet."""
    report = pd.DataFrame({
        'check': ['invalid_chour', 'duplicate_barcode', 'count_mismatch'],
        'severity': ['ERROR', 'ERROR', 'ERROR'],
        'mrccode': ['23', '23', '23'],
        'site': ['Atiak', 'Atiak', 'Atiak'],
        'field': ['chour', 'mosq_barcode', 'numfanoph'],
        'record_count': [1, 2, 3],
        'detail': ['d1', 'd2', 'd3'],
        'clocation': ['', '', ''],
        'hhid': ['h1', 'h2', 'h3'],
        'session_id': ['s1', 's2', 's3'],
    })
    with patch('pandas.read_sql', return_value=pd.DataFrame()):
        result_bytes = _build_validation_details_excel(MagicMock(), report, {})

    wb = load_workbook(_io.BytesIO(result_bytes))
    assert 'invalid_chour' in wb.sheetnames
    assert 'duplicate_barcode' in wb.sheetnames
    assert 'count_mismatch' in wb.sheetnames


def test_build_validation_details_excel_includes_clocation_mismatch_sheet():
    report = pd.DataFrame({
        'check': ['mosquito_clocation_mismatch'],
        'severity': ['ERROR'],
        'mrccode': ['23'],
        'field': ['clocation'],
        'record_count': [2],
        'detail': ['2 mosquito record(s) have a clocation that differs from their parent collection record.'],
        'clocation': ['indoor'],
        'hhid': ['312010595'],
        'session_id': ['sess1'],
        'site': ['Atiak HCIV - Amuru'],
    })
    mismatch_records = pd.DataFrame({
        'session_id': ['sess1'],
        'hhid': ['312010595'],
        'mrccode': ['23'],
        'mosquito_clocation': ['2'],
        'collection_clocation': ['1'],
    })
    with patch('pandas.read_sql', return_value=mismatch_records):
        result_bytes = _build_validation_details_excel(MagicMock(), report, {'23': 'Atiak HCIV - Amuru'})

    wb = load_workbook(_io.BytesIO(result_bytes))
    assert 'mosquito_clocation_mismatch' in wb.sheetnames
