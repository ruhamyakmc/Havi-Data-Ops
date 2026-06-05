from __future__ import annotations

import html as _html
import io
import logging
import smtplib
import ssl
from datetime import date
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
from cryptography.fernet import Fernet
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from stages.base import StageResult

logger = logging.getLogger(__name__)


def _load_smtp_password(ini_path: str, key_path: str) -> str:
    """Read the Fernet-encrypted Password value from an ini-style file."""
    with open(key_path, 'r') as f:
        key = f.read().strip().encode()
    cipher = Fernet(key)

    cfg: dict[str, str] = {}
    with open(ini_path, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            k, _, v = line.partition('=')
            cfg[k.strip()] = v.strip()

    if 'Password' not in cfg:
        raise KeyError("'Password' key not found in SMTP credential file.")

    return cipher.decrypt(cfg['Password'].encode()).decode()


def _query_validation_report(engine) -> pd.DataFrame | None:
    """Query the HAVI validation report if it exists."""
    try:
        return pd.read_sql('SELECT * FROM gold_havi.ds_validation_report', engine)
    except Exception as exc:
        logger.warning("Could not query ds_validation_report: %s", exc)
        return None


def _build_stage_summary(
    results: dict[str, StageResult],
    stages: list[str],
) -> str:
    sep = '-' * 47
    lines = ['Stage Results', sep]
    for name in stages:
        if name not in results:
            lines.append(f'  SKIP  {name:<28}  skipped')
        elif results[name].success:
            rw = results[name].rows_written
            row_str = f'{rw:,} rows' if rw else ''
            lines.append(f'  OK    {name:<28}  {row_str}')
        else:
            lines.append(f'  FAIL  {name:<28}  FAILED')
    lines.append(sep)
    return '\n'.join(lines)


def _build_validation_summary(report_df: pd.DataFrame | None) -> str:
    """Concise summary for the email body; full detail is in the CSV attachment."""
    if report_df is None:
        return 'Validation report unavailable - measures_havi did not run.\n'

    sep = '-' * 47
    lines = ['Validation Issues (see attachment for full detail)', sep]

    for severity in ['ERROR', 'WARNING']:
        subset = report_df[report_df['severity'] == severity]
        if subset.empty:
            continue
        lines.append(f'\n  {severity}S ({len(subset)} issue row(s)):')
        group_cols = [c for c in ['country', 'site', 'mrccode'] if c in subset.columns]
        if group_cols:
            grouped = subset.groupby(group_cols, sort=True, dropna=False)
            for keys, group in grouped:
                if not isinstance(keys, tuple):
                    keys = (keys,)
                header = ' / '.join(str(k) for k in keys if pd.notna(k) and str(k))
                lines.append(f'    {header or "all"}')
                for check, cnt in group.groupby('check').size().items():
                    lines.append(f'      - {check} ({cnt})')
        else:
            for check, cnt in subset.groupby('check').size().items():
                lines.append(f'    - {check} ({cnt})')

    lines.append(sep)
    return '\n'.join(lines)


def _attach_csv(msg: MIMEMultipart, df: pd.DataFrame, filename: str) -> None:
    """Sanitize and attach a DataFrame as a UTF-8 CSV to a MIME message."""
    csv_buffer = io.StringIO()
    safe_df = df.copy()
    for col in safe_df.select_dtypes(include='object').columns:
        safe_df.loc[:, col] = safe_df[col].map(
            lambda v: ("'" + v)
            if isinstance(v, str) and v and v[0] in ('=', '+', '-', '@', '\t', '\r')
            else v
        )
    safe_df.to_csv(csv_buffer, index=False)
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(csv_buffer.getvalue().encode('utf-8-sig'))
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
    msg.attach(part)


def _build_validation_details_excel(engine, report_df: pd.DataFrame, mrc_sites: dict) -> bytes:
    """
    Build an Excel workbook with:
    - 'Summary' sheet: the full validation report
    - One sheet per error/warning check with the actual affected records from gold tables
    """
    hdr_fill  = PatternFill('solid', fgColor='1F4E79')
    err_fill  = PatternFill('solid', fgColor='FFCCCC')
    warn_fill = PatternFill('solid', fgColor='FFF2CC')
    grn_fill  = PatternFill('solid', fgColor='CCFFCC')
    hdr_font  = Font(bold=True, color='FFFFFF')
    bold_font = Font(bold=True, size=12)
    thin      = Side(style='thin')
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr(ws, row, cols):
        for c, name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.border = bdr

    def _row(ws, row, values, fill=None):
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            if fill: cell.fill = fill
            cell.border = bdr

    def _write_df(ws, df, start_row=1, fill=None):
        _hdr(ws, start_row, list(df.columns))
        for ri, r in enumerate(df.itertuples(index=False), start=start_row + 1):
            _row(ws, ri, list(r), fill)
        for i in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

    def _query(sql):
        try:
            return pd.read_sql(sql, engine)
        except Exception as exc:
            logger.warning('Validation detail query failed: %s', exc)
            return pd.DataFrame()

    cloc_label = {'1': 'Outdoor', '2': 'Indoor'}

    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary sheet ---
    ws = wb.create_sheet('Summary')
    ws.cell(1, 1, 'Validation Report').font = bold_font
    display_cols = [c for c in report_df.columns if c not in ('clocation',)]
    _write_df(ws, report_df[display_cols], start_row=2)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions[get_column_letter(len(display_cols))].width = 80

    # --- Per-check detail sheets ---
    checks_present = report_df['check'].unique()

    if 'count_mismatch' in checks_present:
        ws = wb.create_sheet('count_mismatch')
        ws.cell(1, 1, 'Collection sessions where numfanoph ≠ actual mosquito count').font = bold_font
        coll = _query('SELECT session_id, mrccode, hhid, dateofcollection, clocation, numfanoph, nummanoph, numculex FROM gold_havi.ento_collection')
        mosq_c = _query('SELECT session_id, clocation, COUNT(*) as actual_count FROM gold_havi.ento_mosquito GROUP BY session_id, clocation')
        mm = pd.DataFrame()
        mosq_detail = pd.DataFrame()
        if not coll.empty and not mosq_c.empty:
            coll['numfanoph_n'] = pd.to_numeric(coll['numfanoph'], errors='coerce')
            merged = coll.merge(mosq_c, on=['session_id', 'clocation'], how='left')
            merged['actual_count'] = merged['actual_count'].fillna(0).astype(int)
            mm = merged[merged['numfanoph_n'].notna() & (merged['numfanoph_n'] >= 1) & (merged['numfanoph_n'] != merged['actual_count'])].copy()
            mm['site'] = mm['mrccode'].astype(str).map(mrc_sites).fillna('')
            mm['location'] = mm['clocation'].astype(str).map(cloc_label)
            mm['discrepancy'] = (mm['actual_count'] - mm['numfanoph_n'].astype(int)).astype(int)
            out = mm[['site', 'session_id', 'hhid', 'dateofcollection', 'location', 'numfanoph', 'actual_count', 'discrepancy']].rename(columns={
                'site': 'Site', 'session_id': 'Session ID', 'hhid': 'Household ID',
                'dateofcollection': 'Collection Date', 'location': 'Location',
                'numfanoph': 'Declared', 'actual_count': 'Recorded', 'discrepancy': 'Discrepancy',
            })
            _write_df(ws, out, start_row=2, fill=err_fill)
            mosq_detail = _query(
                'SELECT session_id, clocation, mosquito_number, chour, grossspecies, abdstatus, mosq_barcode '
                'FROM gold_havi.ento_mosquito'
            )

        if not mm.empty and not mosq_detail.empty:
            mosq_cols = ['Session ID', 'Household ID', 'Collection Date', 'Location',
                         'Mosquito # (derived)', 'Collection Hour', 'Species', 'Abd. Status', 'Barcode']
            col_widths = [28, 16, 18, 12, 20, 18, 12, 14, 18]
            for site, site_df in mm.groupby('site'):
                sheet_name = (str(site)[:31]) if site else 'Unknown'
                ws_s = wb.create_sheet(title=sheet_name)
                ws_s.cell(1, 1, f'Count Mismatch — {site}').font = bold_font

                ws_s.cell(3, 1, 'SESSION SUMMARY').font = bold_font
                summary_cols = ['Session ID', 'Household ID', 'Collection Date', 'Location',
                                'Declared (numfanoph)', 'Recorded Mosquitoes', 'Discrepancy']
                _hdr(ws_s, 4, summary_cols)
                sum_row = 4
                for _, r in site_df.sort_values('session_id').iterrows():
                    sum_row += 1
                    cloc = cloc_label.get(str(r['clocation']), str(r['clocation']))
                    diff = int(r['discrepancy'])
                    _row(ws_s, sum_row, [
                        r['session_id'], r['hhid'], str(r['dateofcollection']),
                        cloc, int(r['numfanoph_n']), int(r['actual_count']), diff,
                    ], grn_fill if diff > 0 else err_fill)

                offset = sum_row + 2
                ws_s.cell(offset, 1, 'MOSQUITO RECORDS').font = bold_font
                _hdr(ws_s, offset + 1, mosq_cols)
                det_row = offset + 1
                for _, r in site_df.sort_values('session_id').iterrows():
                    cloc = cloc_label.get(str(r['clocation']), str(r['clocation']))
                    rows = mosq_detail[
                        (mosq_detail['session_id'] == r['session_id']) &
                        (mosq_detail['clocation'].astype(str) == str(r['clocation']))
                    ].sort_values('mosquito_number')
                    for _, m in rows.iterrows():
                        det_row += 1
                        _row(ws_s, det_row, [
                            r['session_id'], r['hhid'], str(r['dateofcollection']), cloc,
                            m['mosquito_number'], m['chour'], m['grossspecies'],
                            m['abdstatus'], m['mosq_barcode'],
                        ])
                for i, w in enumerate(col_widths, 1):
                    ws_s.column_dimensions[get_column_letter(i)].width = w

    if 'duplicate_hhid_per_date' in checks_present:
        ws = wb.create_sheet('duplicate_hhid_per_date')
        ws.cell(1, 1, 'Household records sharing the same hhid + date').font = bold_font
        df = _query("""
            SELECT h.mrccode, h.session_id, h.hhid, h.dateofobservation,
                   h.numpeople, h.numsleeprooms, h.numsleepareas, h.numhangbednets,
                   h.starttime, h.lastmod
            FROM gold_havi.hbo_household h
            WHERE EXISTS (
                SELECT 1 FROM gold_havi.hbo_household h2
                WHERE h2.hhid = h.hhid AND h2.dateofobservation = h.dateofobservation
                  AND h2.uniqueid != h.uniqueid
            )
            ORDER BY h.hhid, h.dateofobservation
        """)
        if not df.empty:
            df['site'] = df['mrccode'].astype(str).map(mrc_sites).fillna('')
            cols = ['site', 'session_id', 'hhid', 'dateofobservation', 'numpeople',
                    'numsleeprooms', 'numsleepareas', 'numhangbednets', 'starttime', 'lastmod']
            _write_df(ws, df[cols].rename(columns={'site': 'Site', 'session_id': 'Session ID',
                'hhid': 'Household ID', 'dateofobservation': 'Obs Date'}), start_row=2, fill=err_fill)

    if 'person_count_vs_numpeople' in checks_present:
        ws = wb.create_sheet('person_count_vs_numpeople')
        ws.cell(1, 1, 'Sessions where person record count ≠ declared numpeople').font = bold_font
        df = _query("""
            SELECT h.mrccode, h.session_id, h.hhid, h.dateofobservation,
                   h.numpeople AS declared_numpeople,
                   COUNT(p.uniqueid) AS person_records
            FROM gold_havi.hbo_household h
            LEFT JOIN gold_havi.hbo_person p ON p.session_id = h.session_id
            WHERE h.numpeople ~ '^[0-9]+$'
            GROUP BY h.mrccode, h.session_id, h.hhid, h.dateofobservation, h.numpeople
            HAVING h.numpeople::int != COUNT(p.uniqueid)
            ORDER BY h.mrccode, h.session_id
        """)
        if not df.empty:
            df['site'] = df['mrccode'].astype(str).map(mrc_sites).fillna('')
            df['discrepancy'] = df['person_records'].astype(int) - df['declared_numpeople'].astype(int)
            cols = ['site', 'session_id', 'hhid', 'dateofobservation', 'declared_numpeople', 'person_records', 'discrepancy']
            _write_df(ws, df[cols].rename(columns={'site': 'Site', 'session_id': 'Session ID',
                'hhid': 'Household ID', 'dateofobservation': 'Obs Date',
                'declared_numpeople': 'Declared (numpeople)', 'person_records': 'Actual Records',
                'discrepancy': 'Discrepancy'}), start_row=2, fill=err_fill)

    if 'sleepareas_less_than_sleeprooms' in checks_present:
        ws = wb.create_sheet('sleepareas_lt_sleeprooms')
        ws.cell(1, 1, 'Households where sleeping areas < sleeping rooms').font = bold_font
        df = _query("""
            SELECT mrccode, session_id, hhid, dateofobservation,
                   numsleeprooms, numsleepareas, numhangbednets, numpeople,
                   starttime, lastmod
            FROM gold_havi.hbo_household
            WHERE numsleepareas ~ '^[0-9]+$' AND numsleeprooms ~ '^[0-9]+$'
              AND numsleepareas::int < numsleeprooms::int
            ORDER BY mrccode, session_id
        """)
        if not df.empty:
            df['site'] = df['mrccode'].astype(str).map(mrc_sites).fillna('')
            cols = ['site', 'session_id', 'hhid', 'dateofobservation',
                    'numsleeprooms', 'numsleepareas', 'numhangbednets', 'numpeople', 'starttime', 'lastmod']
            _write_df(ws, df[cols].rename(columns={'site': 'Site', 'session_id': 'Session ID',
                'hhid': 'Household ID', 'dateofobservation': 'Obs Date',
                'numsleeprooms': 'Sleep Rooms', 'numsleepareas': 'Sleep Areas',
                'numhangbednets': 'Bed Nets', 'numpeople': 'Num People'}), start_row=2, fill=err_fill)

    if 'obs_transition_net_out_net' in checks_present:
        ws = wb.create_sheet('obs_transition_net_out_net')
        ws.cell(1, 1, 'Person records with Under-net IN → Near net OUT → Under-net IN transition').font = bold_font
        obs_cols = [f'obs_{h}' for h in [
            '4_5pm','5_6pm','6_7pm','7_8pm','8_9pm','9_10pm',
            '10_11pm','11pm_12am','12_1am','1_2am','2_3am','3_4am','4_5am',
            '5_6am','6_7am','7_8am','8_9am','9_10am',
        ]]
        df = _query('SELECT * FROM gold_havi.hbo_person ORDER BY session_id, individualnum')
        if not df.empty:
            # Find persons with the IN→OUT→IN transition
            present = [c for c in obs_cols if c in df.columns]
            def has_transition(row):
                vals = [pd.to_numeric(row.get(c), errors='coerce') for c in present]
                vals = [v for v in vals if pd.notna(v)]
                for i in range(len(vals) - 2):
                    if vals[i] == 1 and vals[i+1] == 3 and vals[i+2] == 1:
                        return True
                return False
            mask = df.apply(has_transition, axis=1)
            flagged = df[mask].copy()
            if not flagged.empty:
                flagged['site'] = flagged['mrccode'].astype(str).map(mrc_sites).fillna('') if 'mrccode' in flagged.columns else ''
                display = ['session_id', 'hhid', 'dateofobservation', 'individualnum', 'age', 'gender'] + present
                display = [c for c in display if c in flagged.columns]
                _write_df(ws, flagged[display], start_row=2, fill=warn_fill)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _attach_bytes(msg: MIMEMultipart, data: bytes, filename: str) -> None:
    """Attach raw bytes to a MIME message."""
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(data)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
    msg.attach(part)



def _send(
    email_cfg: dict,
    recipients: list[str],
    subject: str,
    plain: str,
    html: str,
    attachment_df: pd.DataFrame | None = None,
    attachment_filename: str | None = None,
    extra_attachments: list[tuple[bytes, str]] | None = None,
    cc: list[str] | None = None,
) -> None:
    """Assemble a multipart email and send it via SMTP with STARTTLS."""
    ini_path = email_cfg['keyfiles']['smtp_ini']
    key_path = email_cfg['keyfiles']['smtp_key']
    username = email_cfg['smtp_username']
    password = _load_smtp_password(ini_path, key_path)

    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From'] = email_cfg['sender']
    msg['To'] = ', '.join(recipients)
    if cc:
        msg['Cc'] = ', '.join(cc)

    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(plain, 'plain'))
    alt.attach(MIMEText(html, 'html'))
    msg.attach(alt)

    if attachment_df is not None:
        filename = attachment_filename or f'havi_validation_{date.today().strftime("%Y-%m-%d")}.csv'
        _attach_csv(msg, attachment_df, filename)

    for data, filename in (extra_attachments or []):
        _attach_bytes(msg, data, filename)

    with smtplib.SMTP(email_cfg['smtp_host'], email_cfg['smtp_port']) as smtp:
        smtp.starttls(context=ssl.create_default_context())
        smtp.login(username, password)
        smtp.sendmail(email_cfg['sender'], recipients + (cc or []), msg.as_string())


def send_pipeline_report(
    results: dict[str, StageResult],
    stages: list[str],
    engine,
    config,
) -> None:
    """
    Send HAVI pipeline status and field data quality emails.

    SMTP errors are caught and logged so notification issues never mask ETL results.
    """
    email_cfg = config.get('email')
    if not email_cfg:
        return

    report_df = _query_validation_report(engine)

    stage_warnings = [w for r in results.values() for w in r.warnings]
    if stage_warnings:
        warnings_df = pd.DataFrame(stage_warnings)
        report_df = (
            pd.concat([warnings_df, report_df], ignore_index=True)
            if report_df is not None else warnings_df
        )

    has_failures = any(not r.success for r in results.values())
    today = date.today().strftime('%d %b %Y')
    stage_section = _build_stage_summary(results, stages)

    pipeline_recipients = email_cfg.get('pipeline_recipients', [])
    if pipeline_recipients:
        status = 'FAILED' if has_failures else 'Run complete'
        subject = f'HAVI Pipeline - {status} ({today})'
        html = f'<pre style="font-family:monospace;font-size:13px">{_html.escape(stage_section)}</pre>'
        try:
            _send(email_cfg, pipeline_recipients, subject, stage_section, html)
            logger.info('Pipeline status email sent to %s.', pipeline_recipients)
        except Exception as exc:
            logger.error('Notifier failed for pipeline recipients: %s', exc)

    field_recipients = email_cfg.get('field_recipients', [])
    if not field_recipients or report_df is None or report_df.empty:
        return
    if 'severity' not in report_df.columns or not report_df['severity'].isin(['ERROR', 'WARNING']).any():
        return
    if date.today().weekday() != 4:  # 4 = Friday
        logger.info('Field quality report skipped — only sent on Fridays.')
        return

    mrc_sites = config.get('mrc_sites') or {}
    today_iso = date.today().strftime('%Y-%m-%d')
    extra_attachments: list[tuple[bytes, str]] = []

    validation_excel = _build_validation_details_excel(engine, report_df, mrc_sites)
    extra_attachments.append((validation_excel, f'havi_validation_report_{today_iso}.xlsx'))

    validation_section = _build_validation_summary(report_df)
    subject = f'HAVI Data Quality - issues found ({today})'
    plain = f'{stage_section}\n\n{validation_section}'
    html = f'<pre style="font-family:monospace;font-size:13px">{_html.escape(plain)}</pre>'
    try:
        _send(
            email_cfg, field_recipients, subject, plain, html,
            extra_attachments=extra_attachments,
        )
        logger.info('Field quality report sent to %s.', field_recipients)
    except Exception as exc:
        logger.error('Notifier failed for field recipients %s: %s', field_recipients, exc)
